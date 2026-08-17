"""UltrON — Reports API (Excel + PDF + CSV generation)"""

import csv
import io
import math
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional


# pyrefly: ignore [missing-import]
from fastapi import APIRouter, Depends, Query, HTTPException
# pyrefly: ignore [missing-import]
from fastapi.responses import StreamingResponse
# pyrefly: ignore [missing-import]
from sqlalchemy.ext.asyncio import AsyncSession
# pyrefly: ignore [missing-import]
from sqlalchemy import select, and_, or_, func
from app.database import get_db
from app.models.telemetry import HistoricalData, Averages, AverageType, DataQuality
from app.models.parameter import Parameter
from app.models.device import Device
from app.models.station import Station
from app.core.security import require_admin
from app.core.logger import get_logger
from app.config import settings
from app.services.report_data import fetch_interval_data, MAX_EXPORT_ROWS

log = get_logger("ultron.reports")
router = APIRouter(prefix="/reports", tags=["Reports"])


def _reports_dir() -> Path:
    d = Path(settings.REPORTS_DIR).resolve()
    d.mkdir(parents=True, exist_ok=True)
    return d


async def _fetch_report_data(
    db: AsyncSession,
    parameter_ids: list[int],
    start: datetime,
    end: datetime,
    avg_type: AverageType,
    step_minutes: int = 0,
) -> tuple[list[Parameter], list]:
    """
    Fetch telemetry data for a report, delegating to the shared
    :func:`fetch_interval_data` so preview and export use identical logic.

    ``avg_type`` and ``step_minutes`` are mapped as follows:

    * ``avg_type != raw`` (Average Reports) → average mode with that avg_type.
    * ``avg_type == raw`` and ``step_minutes > 0`` (Normal Reports) → step mode.
    * ``avg_type == raw`` and ``step_minutes == 0`` → raw mode (capped).
    """
    return await fetch_interval_data(
        db=db,
        parameter_ids=parameter_ids,
        start=start,
        end=end,
        interval_minutes=step_minutes if avg_type == AverageType.raw else 0,
        avg_type=avg_type.value if isinstance(avg_type, AverageType) else str(avg_type),
    )


@router.get("/excel")
async def generate_excel(
    parameter_ids: str = Query(..., description="Comma-separated parameter IDs"),
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    avg_type: AverageType = AverageType.avg_1hr,
    step_minutes: int = Query(0, description="Step interval in minutes for normal (raw) reports"),
    station_name: str = Query("UltrON Station"),
    db: AsyncSession = Depends(get_db),
):
    """Generate an Excel (.xlsx) report for selected parameters and time range."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    if not end:
        end = datetime.utcnow()
    if not start:
        start = end - timedelta(hours=24)

    ids = [int(x) for x in parameter_ids.split(",") if x.strip().isdigit()]
    params, readings = await _fetch_report_data(db, ids, start, end, avg_type, step_minutes)
    param_map = {p.id: p for p in params}

    wb = openpyxl.Workbook()
    # Remove default sheet — we'll add per-day sheets
    wb.remove(wb.active)

    if not readings:
        ws = wb.create_sheet(title="No Data")
        ws.append(["UltrON Industrial Monitoring Report"])
        ws.append([f"Station: {station_name}"])
        ws.append([f"No data available (NA) for the selected range."])
        if step_minutes > 0 and avg_type == AverageType.raw:
            ws.append([f"Report Type: Normal (Step: {step_minutes}min)"])
        else:
            ws.append([f"Report Type: Average ({str(avg_type)})"])
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"])
    else:
        daily_data: dict[str, list] = {}
        for r in readings:
            day_key = r.timestamp.strftime("%Y-%m-%d")
            daily_data.setdefault(day_key, []).append(r)

        header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for day_key in sorted(daily_data.keys()):
            day_readings = daily_data[day_key]
            ws = wb.create_sheet(title=day_key)

            ws.append(["UltrON Industrial Monitoring Report"])
            ws.append([f"Station: {station_name}"])
            ws.append([f"Date: {day_key}"])
            if step_minutes > 0 and avg_type == AverageType.raw:
                report_type_label = f"Normal (Step: {step_minutes}min)"
            else:
                report_type_label = f"Average ({str(avg_type)})"
            ws.append([f"Report Type: {report_type_label}"])
            ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"])
            ws.append([])

            timestamps = sorted(set(r.timestamp for r in day_readings))
            data_by_ts: dict = {ts: {} for ts in timestamps}
            for r in day_readings:
                data_by_ts[r.timestamp][r.parameter_id] = r.value

            header_row = ["Date & Time"] + [
                f"{param_map[pid].tag_name}"
                for pid in ids if pid in param_map
            ]
            ws.append(header_row)
            header_row_idx = ws.max_row
            for col_idx in range(1, len(header_row) + 1):
                cell = ws.cell(row=header_row_idx, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for ts in timestamps:
                row_data = [ts.strftime("%Y/%m/%d %H:%M")]
                for pid in ids:
                    if pid in param_map:
                        val = data_by_ts[ts].get(pid)
                        row_data.append(val if val is not None else "NA")
                ws.append(row_data)

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Stream to client
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"UltrON_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _format_pdf_val(val: Optional[float]) -> str:
    if val is None:
        return "NA"
    s = f"{val:.4f}".rstrip('0').rstrip('.')
    if '.' in s:
        parts = s.split('.')
        if len(parts[1]) < 2:
            return f"{val:.2f}"
        return s
    return f"{val:.2f}"


@router.get("/pdf")
async def generate_pdf(
    parameter_ids: str = Query(..., description="Comma-separated parameter IDs"),
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    avg_type: AverageType = AverageType.avg_1hr,
    step_minutes: int = Query(0, description="Step interval in minutes for normal (raw) reports"),
    station_name: str = Query("AAQMS 1"),
    db: AsyncSession = Depends(get_db),
):
    """Generate a PDF report strictly matching the Sunshine Real-Time Monitoring Report layout (Image 2 reference)."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed")

    if not end:
        end = datetime.utcnow()
    if not start:
        start = end - timedelta(hours=24)

    ids = [int(x) for x in parameter_ids.split(",") if x.strip().isdigit()]
    params, readings = await _fetch_report_data(db, ids, start, end, avg_type, step_minutes)
    param_map = {p.id: p for p in params}
    valid_ids = [pid for pid in ids if pid in param_map]

    # Always PORTRAIT as per Image 2 reference
    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    page_w = 210.0
    margin = 10.0
    printable_w = page_w - (margin * 2.0)  # 190.0 mm

    # Top-Right Sunshine Technologies Logo (Uncovered)
    possible_logos = [
        os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "assets", "sunshine_logo.png")),
        os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "ui_dist", "assets", "sunshine_logo.png")),
        os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "..", "..", "frontend", "public", "assets", "sunshine_logo.png")),
    ]
    logo_found = False
    for lp in possible_logos:
        if os.path.exists(lp):
            pdf.image(lp, x=page_w - margin - 48, y=8, w=48)
            logo_found = True
            break

    # Centered Header Titles (Image 2 Reference Style)
    pdf.set_y(14)
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_text_color(15, 23, 42)
    pdf.cell(0, 8, "Real-Time Monitoring Report", ln=True, align="C")
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(30, 41, 59)
    pdf.cell(0, 6, f"Site Name: {station_name}", ln=True, align="C")

    if step_minutes > 0 and avg_type == AverageType.raw:
        report_type_label = f"Normal Report (Step: {step_minutes}min)"
    else:
        report_type_label = f"Average Report ({str(avg_type.value if hasattr(avg_type, 'value') else avg_type)})"

    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(51, 65, 85)
    pdf.cell(0, 6, f"Report: {report_type_label}", ln=True, align="C")

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 6, f"From Date: {start.strftime('%Y/%m/%d %H:%M:%S')} To Date : {end.strftime('%Y/%m/%d %H:%M:%S')}", ln=True, align="C")
    pdf.ln(6)

    num_params = max(1, len(valid_ids))

    # Data grouping by parameter
    timestamps = sorted(set(r.timestamp for r in readings))
    data_by_ts: dict = {ts: {} for ts in timestamps}
    values_by_param: dict[int, list[float]] = {pid: [] for pid in valid_ids}

    for r in readings:
        data_by_ts[r.timestamp][r.parameter_id] = r.value
        if r.value is not None:
            values_by_param[r.parameter_id].append(r.value)

    # ── Summary Statistics Table (Image 2 Reference Style) ──
    desc_w = 48.0
    summary_param_w = (printable_w - desc_w) / num_params

    pdf.set_fill_color(0, 102, 102)  # Teal #006666
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(desc_w, 7, "Description", border=1, fill=True, align="C")
    for pid in valid_ids:
        p = param_map[pid]
        unit_str = f" - ({p.unit})" if p.unit else ""
        hdr = f"{p.name or p.tag_name}{unit_str}"
        if len(hdr) > 18:
            hdr = hdr[:16] + ".."
        pdf.cell(summary_param_w, 7, hdr, border=1, fill=True, align="C")
    pdf.ln()

    # Calculate summary metrics per parameter
    summary_data = {
        "Prescribed Standards": {},
        "Maximum Data": {},
        "Minimum Data": {},
        "Standard Deviation": {},
        "Maximum Value At Time": {},
        "Minimum Value At Time": {},
        "Data Availability %": {},
    }

    for pid in valid_ids:
        p = param_map[pid]
        p_vals = values_by_param[pid]
        # Prescribed standard string
        min_std = getattr(p, 'alarm_low', None)
        max_std = getattr(p, 'alarm_high', None)
        if min_std is not None or max_std is not None:
            summary_data["Prescribed Standards"][pid] = f"{min_std or 0.0} - {max_std or 100.0}"
        else:
            summary_data["Prescribed Standards"][pid] = "0.0 - 100.0"

        if p_vals:
            max_v = max(p_vals)
            min_v = min(p_vals)
            max_ts_str = "NA"
            min_ts_str = "NA"
            for ts in timestamps:
                if data_by_ts[ts].get(pid) == max_v and max_ts_str == "NA":
                    max_ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")
                if data_by_ts[ts].get(pid) == min_v and min_ts_str == "NA":
                    min_ts_str = ts.strftime("%Y-%m-%d %H:%M:%S")

            # Standard Deviation
            if len(p_vals) > 1:
                mean_v = sum(p_vals) / len(p_vals)
                variance = sum((x - mean_v) ** 2 for x in p_vals) / (len(p_vals) - 1)
                std_v = math.sqrt(variance)
            else:
                std_v = 0.0

            avail_pct = min(100.0, (len(p_vals) / max(1, len(timestamps))) * 100.0)

            summary_data["Maximum Data"][pid] = _format_pdf_val(max_v)
            summary_data["Minimum Data"][pid] = _format_pdf_val(min_v)
            summary_data["Standard Deviation"][pid] = f"{std_v:.2f}"
            summary_data["Maximum Value At Time"][pid] = max_ts_str
            summary_data["Minimum Value At Time"][pid] = min_ts_str
            summary_data["Data Availability %"][pid] = f"{avail_pct:.1f}%"
        else:
            summary_data["Maximum Data"][pid] = "NA"
            summary_data["Minimum Data"][pid] = "NA"
            summary_data["Standard Deviation"][pid] = "0.00"
            summary_data["Maximum Value At Time"][pid] = "NA"
            summary_data["Minimum Value At Time"][pid] = "NA"
            summary_data["Data Availability %"][pid] = "0.0%"

    pdf.set_text_color(15, 23, 42)
    pdf.set_font("Helvetica", "", 7.5)
    for row_name, val_map in summary_data.items():
        pdf.cell(desc_w, 6, row_name, border=1, align="C")
        for pid in valid_ids:
            pdf.cell(summary_param_w, 6, str(val_map[pid]), border=1, align="C")
        pdf.ln()

    pdf.ln(6)

    # ── Main Telemetry Table (Image 2 Reference Style) ──
    sl_w = 14.0
    time_w = 38.0
    main_param_w = (printable_w - sl_w - time_w) / num_params

    # Header Row
    pdf.set_fill_color(0, 102, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(sl_w, 7, "Sl No", border=1, fill=True, align="C")
    pdf.cell(time_w, 7, "Time", border=1, fill=True, align="C")
    for pid in valid_ids:
        p = param_map[pid]
        unit_str = f" - ({p.unit})" if p.unit else ""
        hdr = f"{p.name or p.tag_name}{unit_str}"
        if len(hdr) > 18:
            hdr = hdr[:16] + ".."
        pdf.cell(main_param_w, 7, hdr, border=1, fill=True, align="C")
    pdf.ln()

    # Data Rows
    pdf.set_text_color(15, 23, 42)
    pdf.set_font("Helvetica", "", 8)
    fill = False
    for idx, ts in enumerate(timestamps[:21600], start=1):
        pdf.set_fill_color(244, 248, 248) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.cell(sl_w, 6, str(idx), border=1, fill=True, align="C")
        pdf.cell(time_w, 6, ts.strftime("%Y-%m-%d %H:%M"), border=1, fill=True, align="C")
        for pid in valid_ids:
            val = data_by_ts[ts].get(pid)
            pdf.cell(main_param_w, 6, _format_pdf_val(val), border=1, fill=True, align="C")
        pdf.ln()
        fill = not fill

    buf = io.BytesIO(pdf.output())
    fname = f"Sunshine_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"
    try:
        report_path = _reports_dir() / fname
        with open(report_path, "wb") as f:
            f.write(buf.getvalue())
        log.info(f"PDF saved to {report_path}")
    except Exception as e:
        log.warning(f"Could not save PDF to Reports dir: {e}")
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/export-csv")
async def export_csv(
    parameter_ids: str = Query(..., description="Comma-separated parameter IDs"),
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    avg_type: AverageType = AverageType.avg_1hr,
    step_minutes: int = Query(0, description="Step interval in minutes for normal (raw) reports"),
    station_name: str = Query("UltrON Station"),
    db: AsyncSession = Depends(get_db),
):
    """Generate a CSV report saved to the Reports directory."""
    if not end:
        end = datetime.utcnow()
    if not start:
        start = end - timedelta(hours=24)

    ids = [int(x) for x in parameter_ids.split(",") if x.strip().isdigit()]
    params, readings = await _fetch_report_data(db, ids, start, end, avg_type, step_minutes)
    param_map = {p.id: p for p in params}

    if step_minutes > 0 and avg_type == AverageType.raw:
        rtype = f"Normal (Step: {step_minutes}min)"
    else:
        rtype = f"Average ({str(avg_type)})"

    timestamps = sorted(set(r.timestamp for r in readings))
    data_by_ts: dict = {ts: {} for ts in timestamps}
    for r in readings:
        data_by_ts[r.timestamp][r.parameter_id] = r.value

    headers = ["Date & Time"] + [param_map[pid].tag_name for pid in ids if pid in param_map]

    rows = []
    for ts in timestamps:
        row = [ts.strftime("%Y/%m/%d %H:%M")]
        for pid in ids:
            if pid in param_map:
                val = data_by_ts[ts].get(pid)
                row.append(f"{val:.3f}" if val is not None else "NA")
        rows.append(row)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    csv_content = buf.getvalue()

    fname = f"UltrON_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.csv"
    try:
        report_path = _reports_dir() / fname
        report_path.write_text(csv_content, encoding="utf-8")
        log.info(f"CSV saved to {report_path}")
    except Exception as e:
        log.warning(f"Could not save CSV to Reports dir: {e}")

    return StreamingResponse(
        io.BytesIO(csv_content.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/windrose", dependencies=[Depends(require_admin)])
async def get_windrose_data(
    station_id: int = Query(..., description="Station ID"),
    date_from: str = Query(..., description="Start date YYYY-MM-DD"),
    date_to: str = Query(..., description="End date YYYY-MM-DD"),
    parameter_id: Optional[int] = Query(None, description="Optional parameter ID for pollutionrose"),
):
    """Windrose endpoint — returns empty data set.

    A real windrose requires paired wind-direction + wind-speed parameters
    grouped by compass direction. Without both, any data returned is
    misleading. The frontend falls back to sample data when datasets=[].
    """
    log.info("Windrose endpoint called without paired direction data — returning empty (frontend will fall back to sample)")
    directions = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE", "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]
    return {
        "station_id": station_id,
        "date_from": date_from,
        "date_to": date_to,
        "labels": directions,
        "datasets": [],
    }


# ── Analytical Reports ────────────────────────────────────────────────────────


async def _fetch_param_values(
    db: AsyncSession, param_id: int, station_id: int, start: str, end: str, limit: int = 5000
) -> list[float]:
    try:
        start_dt = datetime.strptime(start[:10], "%Y-%m-%d")
        end_dt = datetime.strptime(end[:10], "%Y-%m-%d") + timedelta(days=1)
    except ValueError:
        return []
    stmt = select(HistoricalData.value).join(Parameter, HistoricalData.parameter_id == Parameter.id).join(
        Device, Parameter.device_id == Device.id
    ).where(
        Device.station_id == station_id,
        HistoricalData.parameter_id == param_id,
        HistoricalData.timestamp >= start_dt,
        HistoricalData.timestamp <= end_dt,
        HistoricalData.quality.in_((DataQuality.good, DataQuality.out_of_range)),
    ).limit(limit)
    res = await db.execute(stmt)
    return [r[0] for r in res.all() if r[0] is not None]


@router.get("/histogram", dependencies=[Depends(require_admin)])
async def get_histogram(
    station: int = Query(...),
    parameter: int = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    bins: int = Query(10, ge=2, le=100),
    db: AsyncSession = Depends(get_db),
):
    """Return frequency distribution (histogram bins) for a parameter."""
    values = await _fetch_param_values(db, parameter, station, start, end)
    if not values:
        return {"bins": [], "total": 0}

    min_v, max_v = min(values), max(values)
    if max_v == min_v:
        return {"bins": [{"range": f"{min_v}", "count": len(values)}], "total": len(values)}

    bin_w = (max_v - min_v) / bins
    bin_counts = [0] * bins
    for v in values:
        idx = min(int((v - min_v) / bin_w), bins - 1)
        bin_counts[idx] += 1

    result_bins = [
        {"range": f"{min_v + i * bin_w:.2f}–{min_v + (i + 1) * bin_w:.2f}", "count": c}
        for i, c in enumerate(bin_counts)
    ]
    return {"bins": result_bins, "total": len(values)}


@router.get("/percentile", dependencies=[Depends(require_admin)])
async def get_percentile(
    station: int = Query(...),
    parameter: int = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Return percentile values (P10–P99) for a parameter."""
    values = await _fetch_param_values(db, parameter, station, start, end)
    if not values:
        return {"p10": None, "p25": None, "p50": None, "p75": None, "p90": None, "p95": None, "p99": None}

    values.sort()
    n = len(values)

    def percentile(p):
        k = (p / 100.0) * (n - 1)
        f = math.floor(k)
        c = math.ceil(k)
        if f == c:
            return round(values[int(k)], 2)
        d0 = values[f] * (c - k)
        d1 = values[c] * (k - f)
        return round(d0 + d1, 2)

    return {
        "p10": percentile(10), "p25": percentile(25), "p50": percentile(50),
        "p75": percentile(75), "p90": percentile(90), "p95": percentile(95), "p99": percentile(99),
    }


@router.get("/scatter", dependencies=[Depends(require_admin)])
async def get_scatter(
    x_param: int = Query(...),
    y_param: int = Query(...),
    station: int = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    limit: int = Query(500, le=5000),
    db: AsyncSession = Depends(get_db),
):
    """Return scatter plot points for two time-aligned parameters."""
    try:
        start_dt = datetime.strptime(start[:10], "%Y-%m-%d")
        end_dt = datetime.strptime(end[:10], "%Y-%m-%d") + timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    stmt = select(HistoricalData.timestamp, HistoricalData.value).where(
        HistoricalData.parameter_id == x_param,
        HistoricalData.timestamp >= start_dt,
        HistoricalData.timestamp <= end_dt,
        HistoricalData.quality.in_((DataQuality.good, DataQuality.out_of_range)),
    ).order_by(HistoricalData.timestamp).limit(limit)
    x_res = await db.execute(stmt)
    x_data = {r[0].isoformat(): r[1] for r in x_res.all() if r[1] is not None}

    stmt = select(HistoricalData.timestamp, HistoricalData.value).where(
        HistoricalData.parameter_id == y_param,
        HistoricalData.timestamp >= start_dt,
        HistoricalData.timestamp <= end_dt,
        HistoricalData.quality.in_((DataQuality.good, DataQuality.out_of_range)),
    ).order_by(HistoricalData.timestamp).limit(limit)
    y_res = await db.execute(stmt)
    y_map = {r[0].isoformat(): r[1] for r in y_res.all() if r[1] is not None}

    common_ts = sorted(set(x_data.keys()) & set(y_map.keys()))
    points = [{"x": round(x_data[ts], 2), "y": round(y_map[ts], 2)} for ts in common_ts]
    return {"points": points}


@router.get("/uptime", dependencies=[Depends(require_admin)])
async def get_uptime(
    station: int = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Return daily data availability for a station."""
    try:
        start_dt = datetime.strptime(start[:10], "%Y-%m-%d")
        end_dt = datetime.strptime(end[:10], "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    param_ids_res = await db.execute(
        select(Parameter.id).join(Device).where(Device.station_id == station, Parameter.is_active == True)
    )
    param_ids = [r[0] for r in param_ids_res.all()]
    if not param_ids:
        return {"days": []}

    days = []
    cursor = start_dt
    while cursor <= end_dt:
        day_start = cursor
        day_end = cursor + timedelta(days=1)
        count_res = await db.execute(
            select(func.count(HistoricalData.id)).where(
                HistoricalData.parameter_id.in_(param_ids),
                HistoricalData.timestamp >= day_start,
                HistoricalData.timestamp <= day_end,
            )
        )
        total_points = count_res.scalar() or 0

        valid_res = await db.execute(
            select(func.count(HistoricalData.id)).where(
                HistoricalData.parameter_id.in_(param_ids),
                HistoricalData.timestamp >= day_start,
                HistoricalData.timestamp <= day_end,
                HistoricalData.quality.in_((DataQuality.good, DataQuality.out_of_range)),
            )
        )
        valid_points = valid_res.scalar() or 0

        days.append({
            "date": cursor.strftime("%d-%m-%Y"),
            "total_points": total_points,
            "valid_points": valid_points,
            "availability_pct": round((valid_points / total_points) * 100, 1) if total_points > 0 else 0,
        })
        cursor += timedelta(days=1)
    return {"days": days}


@router.get("/shift", dependencies=[Depends(require_admin)])
async def get_shift(
    station: int = Query(...),
    start: str = Query(...),
    end: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Return per-shift summary statistics."""
    try:
        start_dt = datetime.strptime(start[:10], "%Y-%m-%d")
        end_dt = datetime.strptime(end[:10], "%Y-%m-%d") + timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    param_ids_res = await db.execute(
        select(Parameter.id).join(Device).where(Device.station_id == station, Parameter.is_active == True)
    )
    param_ids = [r[0] for r in param_ids_res.all()]
    if not param_ids:
        return {"shifts": []}

    shifts_config = [
        {"name": "Morning (06-14)", "start_h": 6, "end_h": 14},
        {"name": "Evening (14-22)", "start_h": 14, "end_h": 22},
        {"name": "Night (22-06)", "start_h": 22, "end_h": 6},
    ]

    shifts = []
    for sc in shifts_config:
        if sc["start_h"] < sc["end_h"]:
            hour_condition = and_(
                func.extract("hour", HistoricalData.timestamp) >= sc["start_h"],
                func.extract("hour", HistoricalData.timestamp) < sc["end_h"],
            )
        else:
            hour_condition = or_(
                func.extract("hour", HistoricalData.timestamp) >= sc["start_h"],
                func.extract("hour", HistoricalData.timestamp) < sc["end_h"],
            )

        agg_res = await db.execute(
            select(
                func.avg(HistoricalData.value).label("avg"),
                func.min(HistoricalData.value).label("min"),
                func.max(HistoricalData.value).label("max"),
            ).where(
                HistoricalData.parameter_id.in_(param_ids),
                HistoricalData.timestamp >= start_dt,
                HistoricalData.timestamp <= end_dt,
                HistoricalData.quality.in_((DataQuality.good, DataQuality.out_of_range)),
                hour_condition,
            )
        )
        row = agg_res.one()
        shifts.append({
            "name": sc["name"],
            "avg": round(row.avg, 1) if row.avg else 0,
            "min": round(row.min, 1) if row.min else 0,
            "max": round(row.max, 1) if row.max else 0,
        })
    return {"shifts": shifts}


@router.get("/fortnight", dependencies=[Depends(require_admin)])
async def get_fortnight(
    station: int = Query(...),
    month: str = Query(...),
    year: str = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """Return 15-day block summaries for a given month."""
    try:
        m = int(month)
        y = int(year)
    except ValueError:
        raise HTTPException(status_code=400, detail="month and year must be integers.")

    param_ids_res = await db.execute(
        select(Parameter.id).join(Device).where(Device.station_id == station, Parameter.is_active == True)
    )
    param_ids = [r[0] for r in param_ids_res.all()]

    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    m_label = month_names[m - 1] if 1 <= m <= 12 else str(m)

    blocks = []
    for half, day_start in [(1, 1), (2, 16)]:
        if m == 12:
            next_m, next_y = 1, y + 1
        else:
            next_m, next_y = m + 1, y
        block_start = datetime(y, m, day_start)
        if day_start == 16:
            block_end = datetime(next_y, next_m, 1)
        else:
            block_end = datetime(y, m, 15, 23, 59, 59)

        agg = {"avg": None, "availability_pct": 0}
        if param_ids:
            agg_res = await db.execute(
                select(func.avg(HistoricalData.value)).where(
                    HistoricalData.parameter_id.in_(param_ids),
                    HistoricalData.timestamp >= block_start,
                    HistoricalData.timestamp < block_end,
                    HistoricalData.quality.in_((DataQuality.good, DataQuality.out_of_range)),
                )
            )
            avg_val = agg_res.scalar()
            agg["avg"] = round(avg_val, 1) if avg_val else None

            total_res = await db.execute(
                select(func.count(HistoricalData.id)).where(
                    HistoricalData.parameter_id.in_(param_ids),
                    HistoricalData.timestamp >= block_start,
                    HistoricalData.timestamp < block_end,
                )
            )
            total_pts = total_res.scalar() or 0
            valid_res = await db.execute(
                select(func.count(HistoricalData.id)).where(
                    HistoricalData.parameter_id.in_(param_ids),
                    HistoricalData.timestamp >= block_start,
                    HistoricalData.timestamp < block_end,
                    HistoricalData.quality.in_((DataQuality.good, DataQuality.out_of_range)),
                )
            )
            valid_pts = valid_res.scalar() or 0
            agg["availability_pct"] = round((valid_pts / total_pts) * 100, 1) if total_pts > 0 else 0

        blocks.append({
            "label": f"{'1-15' if half == 1 else '16-' + m_label[-1]} {m_label}",
            "availability_pct": agg["availability_pct"],
            "parameters": {"avg_value": agg["avg"]},
        })
    return {"blocks": blocks}
