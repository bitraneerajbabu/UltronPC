"""UltrON — Reports API (Excel + PDF generation)"""

import os
import io
from datetime import datetime, timedelta
from typing import Optional


# pyrefly: ignore [missing-import]
from fastapi import APIRouter, Depends, Query, HTTPException
# pyrefly: ignore [missing-import]
from fastapi.responses import StreamingResponse
# pyrefly: ignore [missing-import]
from sqlalchemy.ext.asyncio import AsyncSession
# pyrefly: ignore [missing-import]
from sqlalchemy import select, and_
from app.database import get_db
from app.models.telemetry import HistoricalData, Averages, AverageType, DataQuality
from app.models.parameter import Parameter
from app.models.station import Station
from app.config import settings
from app.core.logger import get_logger

log = get_logger("ultron.reports")
router = APIRouter(prefix="/reports", tags=["Reports"])


async def _fetch_report_data(
    db: AsyncSession,
    parameter_ids: list[int],
    start: datetime,
    end: datetime,
    avg_type: AverageType,
) -> tuple[list[Parameter], list]:
    param_result = await db.execute(
        select(Parameter).where(Parameter.id.in_(parameter_ids))
    )
    params = param_result.scalars().all()

    if avg_type == AverageType.raw:
        model = HistoricalData
    else:
        model = Averages

    query = select(model).where(
        and_(
            model.parameter_id.in_(parameter_ids),
            model.timestamp >= start,
            model.timestamp <= end,
        )
    )
    if avg_type != AverageType.raw:
        query = query.where(model.avg_type == avg_type)

    tel_result = await db.execute(
        query.order_by(model.timestamp)
    )
    readings = tel_result.scalars().all()
    return params, readings


@router.get("/excel")
async def generate_excel(
    parameter_ids: str = Query(..., description="Comma-separated parameter IDs"),
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    avg_type: AverageType = AverageType.avg_1hr,
    station_name: str = Query("UltrON Station"),
    db: AsyncSession = Depends(get_db),
):
    """Generate an Excel (.xlsx) report for selected parameters and time range."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    if not end:
        end = datetime.utcnow()
    if not start:
        start = end - timedelta(hours=24)

    ids = [int(x) for x in parameter_ids.split(",") if x.strip().isdigit()]
    params, readings = await _fetch_report_data(db, ids, start, end, avg_type)
    param_map = {p.id: p for p in params}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "UltrON Report"

    # ─── Header Styling ───────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    ws.append(["UltrON Industrial Monitoring Report"])
    ws.append([f"Station: {station_name}"])
    ws.append([f"Period: {start.strftime('%Y-%m-%d %H:%M')} to {end.strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"Average Type: {str(avg_type)}"])
    ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"])
    ws.append([])

    # Build pivot: rows = timestamps, cols = parameters
    timestamps = sorted(set(r.timestamp for r in readings))
    data_by_ts: dict = {ts: {} for ts in timestamps}
    for r in readings:
        data_by_ts[r.timestamp][r.parameter_id] = r.value

    # Column headers
    header_row = ["Timestamp"] + [
        f"{param_map[pid].tag_name} ({param_map[pid].unit or '-'})"
        for pid in ids if pid in param_map
    ]
    ws.append(header_row)
    header_row_idx = ws.max_row
    for col_idx, _ in enumerate(header_row, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for ts in timestamps:
        row_data = [ts.strftime("%Y-%m-%d %H:%M:%S")]
        for pid in ids:
            if pid in param_map:
                row_data.append(data_by_ts[ts].get(pid))
        ws.append(row_data)

    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Stream to client
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"UltrON_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/pdf")
async def generate_pdf(
    parameter_ids: str = Query(..., description="Comma-separated parameter IDs"),
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    avg_type: AverageType = AverageType.avg_1hr,
    station_name: str = Query("UltrON Station"),
    db: AsyncSession = Depends(get_db),
):
    """Generate a PDF report."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed")

    if not end:
        end = datetime.utcnow()
    if not start:
        start = end - timedelta(hours=24)

    ids = [int(x) for x in parameter_ids.split(",") if x.strip().isdigit()]
    params, readings = await _fetch_report_data(db, ids, start, end, avg_type)
    param_map = {p.id: p for p in params}

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_fill_color(0, 102, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, "UltrON Industrial Monitoring Report", fill=True, ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, f"Station: {station_name}  |  Period: {start.strftime('%Y-%m-%d %H:%M')} to {end.strftime('%Y-%m-%d %H:%M')}", ln=True)
    pdf.cell(0, 7, f"Avg Type: {avg_type}  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", ln=True)
    pdf.ln(6)

    # Table header
    col_w = 38
    pdf.set_fill_color(0, 102, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(48, 8, "Timestamp", border=1, fill=True)
    for pid in ids:
        if pid in param_map:
            p = param_map[pid]
            pdf.cell(col_w, 8, f"{p.tag_name} ({p.unit or '-'})", border=1, fill=True)
    pdf.ln()

    # Data rows
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    timestamps = sorted(set(r.timestamp for r in readings))
    data_by_ts: dict = {ts: {} for ts in timestamps}
    for r in readings:
        data_by_ts[r.timestamp][r.parameter_id] = r.value

    fill = False
    for ts in timestamps[:500]:   # cap PDF at 500 rows
        pdf.set_fill_color(240, 248, 248) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.cell(48, 7, ts.strftime("%Y-%m-%d %H:%M:%S"), border=1, fill=True)
        for pid in ids:
            val = data_by_ts[ts].get(pid)
            pdf.cell(col_w, 7, f"{val:.3f}" if val is not None else "---", border=1, fill=True)
        pdf.ln()
        fill = not fill

    buf = io.BytesIO(pdf.output())
    fname = f"UltrON_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
