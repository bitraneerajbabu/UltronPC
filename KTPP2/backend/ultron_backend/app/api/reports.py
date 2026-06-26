"""UltrON — Reports API (Excel + PDF generation)"""

import os
import io
import random
from datetime import datetime, timedelta
from typing import Optional


# pyrefly: ignore [missing-import]
from fastapi import APIRouter, Depends, Query, HTTPException
# pyrefly: ignore [missing-import]
from fastapi.responses import StreamingResponse
# pyrefly: ignore [missing-import]
from sqlalchemy.ext.asyncio import AsyncSession
# pyrefly: ignore [missing-import]
from sqlalchemy import select, and_
from app.database import get_db
from app.models.telemetry import HistoricalData, Averages, AverageType, DataQuality
from app.models.parameter import Parameter
from app.models.station import Station
from app.config import settings
from app.core.logger import get_logger

log = get_logger("ultron.reports")
router = APIRouter(prefix="/reports", tags=["Reports"])


async def _fetch_report_data(
    db: AsyncSession,
    parameter_ids: list[int],
    start: datetime,
    end: datetime,
    avg_type: AverageType,
    step_minutes: int = 0,
) -> tuple[list[Parameter], list]:
    param_result = await db.execute(
        select(Parameter).where(Parameter.id.in_(parameter_ids))
    )
    params = param_result.scalars().all()

    if avg_type == AverageType.raw:
        model = HistoricalData
    else:
        model = Averages

    query = select(model).where(
        and_(
            model.parameter_id.in_(parameter_ids),
            model.timestamp >= start,
            model.timestamp <= end,
        )
    )
    if avg_type != AverageType.raw:
        query = query.where(model.avg_type == avg_type)

    tel_result = await db.execute(
        query.order_by(model.timestamp)
    )
    readings = list(tel_result.scalars().all())

    # Step filtering: keep one point per interval for normal (raw) reports
    if step_minutes > 0 and avg_type == AverageType.raw:
        seen = {}
        filtered = []
        for r in readings:
            bucket = r.timestamp.replace(second=0, microsecond=0)
            bucket_key = bucket.minute // step_minutes
            bucket_ts = bucket.replace(minute=bucket_key * step_minutes)
            key = (r.parameter_id, bucket_ts)
            if key not in seen:
                seen[key] = True
                filtered.append(r)
        readings = filtered

    return params, readings


@router.get("/excel")
async def generate_excel(
    parameter_ids: str = Query(..., description="Comma-separated parameter IDs"),
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    avg_type: AverageType = AverageType.avg_1hr,
    step_minutes: int = Query(0, description="Step interval in minutes for normal (raw) reports"),
    station_name: str = Query("UltrON Station"),
    db: AsyncSession = Depends(get_db),
):
    """Generate an Excel (.xlsx) report for selected parameters and time range."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    if not end:
        end = datetime.utcnow()
    if not start:
        start = end - timedelta(hours=24)

    ids = [int(x) for x in parameter_ids.split(",") if x.strip().isdigit()]
    params, readings = await _fetch_report_data(db, ids, start, end, avg_type, step_minutes)
    param_map = {p.id: p for p in params}

    wb = openpyxl.Workbook()
    # Remove default sheet — we'll add per-day sheets
    wb.remove(wb.active)

    # Group readings by date
    daily_data: dict[str, list] = {}
    for r in readings:
        day_key = r.timestamp.strftime("%Y-%m-%d")
        daily_data.setdefault(day_key, []).append(r)

    header_fill = PatternFill(start_color="006666", end_color="006666", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for day_key in sorted(daily_data.keys()):
        day_readings = daily_data[day_key]
        ws = wb.create_sheet(title=day_key)

        # Summary header
        ws.append(["UltrON Industrial Monitoring Report"])
        ws.append([f"Station: {station_name}"])
        ws.append([f"Date: {day_key}"])
        if step_minutes > 0 and avg_type == AverageType.raw:
            report_type_label = f"Normal (Step: {step_minutes}min)"
        else:
            report_type_label = f"Average ({str(avg_type)})"
        ws.append([f"Report Type: {report_type_label}"])
        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC"])
        ws.append([])

        # Pivot: rows = timestamps, cols = parameters
        timestamps = sorted(set(r.timestamp for r in day_readings))
        data_by_ts: dict = {ts: {} for ts in timestamps}
        for r in day_readings:
            data_by_ts[r.timestamp][r.parameter_id] = r.value

        # Column headers
        header_row = ["Date & Time"] + [
            f"{param_map[pid].tag_name}"
            for pid in ids if pid in param_map
        ]
        ws.append(header_row)
        header_row_idx = ws.max_row
        for col_idx in range(1, len(header_row) + 1):
            cell = ws.cell(row=header_row_idx, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for ts in timestamps:
            row_data = [ts.strftime("%Y/%m/%d %H:%M")]
            for pid in ids:
                if pid in param_map:
                    val = data_by_ts[ts].get(pid)
                    row_data.append(val if val is not None else "NA")
            ws.append(row_data)

        # Auto-fit columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

    # Stream to client
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"UltrON_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/pdf")
async def generate_pdf(
    parameter_ids: str = Query(..., description="Comma-separated parameter IDs"),
    start: Optional[datetime] = None,
    end: Optional[datetime] = None,
    avg_type: AverageType = AverageType.avg_1hr,
    step_minutes: int = Query(0, description="Step interval in minutes for normal (raw) reports"),
    station_name: str = Query("UltrON Station"),
    db: AsyncSession = Depends(get_db),
):
    """Generate a PDF report."""
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 not installed")

    if not end:
        end = datetime.utcnow()
    if not start:
        start = end - timedelta(hours=24)

    ids = [int(x) for x in parameter_ids.split(",") if x.strip().isdigit()]
    params, readings = await _fetch_report_data(db, ids, start, end, avg_type, step_minutes)
    param_map = {p.id: p for p in params}

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_fill_color(0, 102, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, "UltrON Industrial Monitoring Report", fill=True, ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 7, f"Station: {station_name}  |  Period: {start.strftime('%Y-%m-%d %H:%M')} to {end.strftime('%Y-%m-%d %H:%M')}", ln=True)
    if step_minutes > 0 and avg_type == AverageType.raw:
        report_type_label = f"Normal (Step: {step_minutes}min)"
    else:
        report_type_label = f"Average ({str(avg_type)})"
    pdf.cell(0, 7, f"Type: {report_type_label}  |  Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC", ln=True)
    pdf.ln(6)

    # Table header
    col_w = 38
    pdf.set_fill_color(0, 102, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(48, 8, "Date & Time", border=1, fill=True)
    for pid in ids:
        if pid in param_map:
            p = param_map[pid]
            pdf.cell(col_w, 8, f"{p.tag_name}", border=1, fill=True)
    pdf.ln()

    # Data rows
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    timestamps = sorted(set(r.timestamp for r in readings))
    data_by_ts: dict = {ts: {} for ts in timestamps}
    for r in readings:
        data_by_ts[r.timestamp][r.parameter_id] = r.value

    fill = False
    for ts in timestamps[:21600]:   # cap at 15 days (21600 rows)
        pdf.set_fill_color(240, 248, 248) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.cell(48, 7, ts.strftime("%Y/%m/%d %H:%M"), border=1, fill=True)
        for pid in ids:
            val = data_by_ts[ts].get(pid)
            pdf.cell(col_w, 7, f"{val:.3f}" if val is not None else "NA", border=1, fill=True)
        pdf.ln()
        fill = not fill

    buf = io.BytesIO(pdf.output())
    fname = f"UltrON_Report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/windrose")
async def get_windrose_data(
    station_id: int = Query(..., description="Station ID"),
    date_from: str = Query(..., description="Start date YYYY-MM-DD"),
    date_to: str = Query(..., description="End date YYYY-MM-DD"),
    parameter_id: Optional[int] = Query(None, description="Optional parameter ID for pollutionrose"),
):
    """Stub endpoint returning sample windrose / pollutionrose data.

    The full implementation will compute wind direction and speed distributions
    from telemetry data stored in HistoricalData / Averages tables.

    Returns radar-chart-friendly data with 16 compass directions and speed bands
    (or pollutant concentrations for pollutionrose mode).
    """
    # 16 wind directions (compass)
    directions = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE", "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]

    if parameter_id:
        # Pollutionrose: return sample pollutant concentrations per direction
        bands = [
            {"label": "PM2.5", "data": [round(random.uniform(10, 80), 1) for _ in range(16)]},
            {"label": "PM10",  "data": [round(random.uniform(20, 150), 1) for _ in range(16)]},
            {"label": "NO2",   "data": [round(random.uniform(5, 60), 1) for _ in range(16)]},
            {"label": "SO2",   "data": [round(random.uniform(2, 30), 1) for _ in range(16)]},
            {"label": "O3",    "data": [round(random.uniform(10, 100), 1) for _ in range(16)]},
            {"label": "CO",    "data": [round(random.uniform(0.5, 5), 2) for _ in range(16)]},
        ]
    else:
        # Windrose: return wind speed bands per direction
        bands = [
            {"label": "0.5–2 m/s", "data": [round(random.uniform(0, 20), 1) for _ in range(16)]},
            {"label": "2–4 m/s",   "data": [round(random.uniform(0, 30), 1) for _ in range(16)]},
            {"label": "4–6 m/s",   "data": [round(random.uniform(0, 25), 1) for _ in range(16)]},
            {"label": "6–8 m/s",   "data": [round(random.uniform(0, 15), 1) for _ in range(16)]},
            {"label": ">8 m/s",    "data": [round(random.uniform(0, 8), 1) for _ in range(16)]},
        ]

    return {
        "station_id": station_id,
        "date_from": date_from,
        "date_to": date_to,
        "labels": directions,
        "datasets": bands,
    }


# ── Analytical Report Stubs ───────────────────────────────────────────────────


@router.get("/histogram")
async def get_histogram(
    station: str = Query(...),
    parameter: str = Query(...),
    start: str = Query(...),
    end: str = Query(...),
):
    """Return frequency distribution (histogram bins) for a parameter."""
    bins = []
    total = 0
    for i in range(0, 100, 10):
        count = random.randint(1, 40)
        bins.append({"range": f"{i}-{i+10}", "count": count})
        total += count
    return {"bins": bins, "total": total}


@router.get("/percentile")
async def get_percentile(
    station: str = Query(...),
    parameter: str = Query(...),
    start: str = Query(...),
    end: str = Query(...),
):
    """Return percentile values (P10–P99) for a parameter."""
    return {
        "p10": round(random.uniform(5, 20), 1),
        "p25": round(random.uniform(20, 35), 1),
        "p50": round(random.uniform(35, 50), 1),
        "p75": round(random.uniform(55, 70), 1),
        "p90": round(random.uniform(75, 88), 1),
        "p95": round(random.uniform(88, 95), 1),
        "p99": round(random.uniform(95, 99), 1),
    }


@router.get("/scatter")
async def get_scatter(
    x_param: str = Query(...),
    y_param: str = Query(...),
    station: str = Query(...),
    start: str = Query(...),
    end: str = Query(...),
):
    """Return scatter plot points for two parameters."""
    points = []
    for _ in range(80):
        x = round(random.uniform(0, 100), 1)
        y = round(x * random.uniform(0.5, 0.9) + random.uniform(0, 15), 1)
        points.append({"x": x, "y": y})
    return {"points": points}


@router.get("/uptime")
async def get_uptime(
    station: str = Query(...),
    start: str = Query(...),
    end: str = Query(...),
):
    """Return daily data availability for a station."""
    try:
        start_dt = datetime.strptime(start[:10], "%Y-%m-%d")
        end_dt = datetime.strptime(end[:10], "%Y-%m-%d")
    except ValueError:
        start_dt = datetime.utcnow() - timedelta(days=7)
        end_dt = datetime.utcnow()

    days = []
    cursor = start_dt
    while cursor <= end_dt:
        total = 1440
        valid = random.randint(int(total * 0.82), total)
        days.append({
            "date": cursor.strftime("%d-%m-%Y"),
            "total_points": total,
            "valid_points": valid,
            "availability_pct": round((valid / total) * 100, 1),
        })
        cursor += timedelta(days=1)
    return {"days": days}


@router.get("/shift")
async def get_shift(
    station: str = Query(...),
    start: str = Query(...),
    end: str = Query(...),
):
    """Return per-shift summary statistics."""
    shifts = [
        {
            "name": "Morning (06-14)",
            "avg": round(random.uniform(30, 50), 1),
            "min": round(random.uniform(5, 15), 1),
            "max": round(random.uniform(60, 80), 1),
        },
        {
            "name": "Evening (14-22)",
            "avg": round(random.uniform(25, 45), 1),
            "min": round(random.uniform(5, 15), 1),
            "max": round(random.uniform(55, 75), 1),
        },
        {
            "name": "Night (22-06)",
            "avg": round(random.uniform(20, 40), 1),
            "min": round(random.uniform(3, 12), 1),
            "max": round(random.uniform(45, 65), 1),
        },
    ]
    return {"shifts": shifts}


@router.get("/fortnight")
async def get_fortnight(
    station: str = Query(...),
    month: str = Query(...),
    year: str = Query(...),
):
    """Return 15-day block summaries for a given month."""
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    try:
        m_idx = int(month) - 1
        m_label = month_names[m_idx] if 0 <= m_idx < 12 else month
    except ValueError:
        m_label = month

    blocks = [
        {
            "label": f"1-15 {m_label}",
            "availability_pct": round(random.uniform(85, 99), 1),
            "parameters": {
                "PM2.5": round(random.uniform(15, 60), 1),
                "PM10": round(random.uniform(30, 120), 1),
                "NO2": round(random.uniform(10, 40), 1),
            },
        },
        {
            "label": f"16-{m_label[-1]} {m_label}",
            "availability_pct": round(random.uniform(85, 99), 1),
            "parameters": {
                "PM2.5": round(random.uniform(15, 60), 1),
                "PM10": round(random.uniform(30, 120), 1),
                "NO2": round(random.uniform(10, 40), 1),
            },
        },
    ]
    return {"blocks": blocks}
